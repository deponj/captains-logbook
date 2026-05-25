"""Convert Captain's Logbook.xlsx → sample-data.json for the PWA.

Output schema matches Settings → Import JSON expectations.
Run: python import_excel.py <input.xlsx> <output.json>
"""
import sys, json, datetime
import openpyxl

XLSX = sys.argv[1]
OUT  = sys.argv[2]

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ── Airports ────────────────────────────────────────────────
airports = []
if 'Airports' in wb.sheetnames:
    ws = wb['Airports']
    # find header row
    headers = None
    for r in range(1, 6):
        row = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
        if any(str(v).lower().startswith('iata') if v else False for v in row):
            headers = [str(v).strip().lower() if v else '' for v in row]
            data_start = r + 1
            break
    if not headers:
        # assume first row is header
        headers = [str(ws.cell(1,c).value).strip().lower() if ws.cell(1,c).value else '' for c in range(1, ws.max_column+1)]
        data_start = 2
    def col(name):
        for i,h in enumerate(headers):
            if name in h: return i
        return -1
    iata_i = col('iata') if col('iata') >= 0 else col('icao')
    lat_i  = col('lat')
    lon_i  = col('lon') if col('lon') >= 0 else col('lng')
    name_i = col('name') if col('name') >= 0 else col('airport')
    for r in range(data_start, ws.max_row+1):
        row = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
        iata = row[iata_i] if iata_i >= 0 else None
        if not iata: continue
        try:
            lat = float(row[lat_i]); lon = float(row[lon_i])
        except (TypeError, ValueError):
            continue
        airports.append({
            'iata': str(iata).strip().upper(),
            'lat': lat, 'lon': lon,
            'name': str(row[name_i]).strip() if name_i >= 0 and row[name_i] else '',
        })

# ── Flights ─────────────────────────────────────────────────
ws = wb['Database']
# header at row 3 based on inspection
HEADER_ROW = 3
hdr = {str(ws.cell(HEADER_ROW,c).value).strip(): c for c in range(1, ws.max_column+1) if ws.cell(HEADER_ROW,c).value}
def col(name_substr):
    for k,v in hdr.items():
        if name_substr.lower() in k.lower():
            return v
    return None

C = {
    'date':    col('Date'),
    'flt':     col('Flt No'),
    'reg':     col('Reg'),
    'dep':     col('DEP'),
    'dest':    col('DEST'),
    'off':     col('Off-Block'),
    'ab':      col('A/B'),
    'td':      col('T/D'),
    'on':      col('On-Block'),
    'isLD':    col('isL/D'),
    'isPIC':   col('isPIC'),
    'crew1':   col('Crew1'),
    'crew2':   col('Crew2'),
    'crew3':   col('Crew3'),
    'block':   col('Block'),
    'flight':  col('Flight'),
    'night':   col('Night'),
    'ldType':  col('L/D Type'),
    'remarks': col('Remarks'),
}

def to_iso_dt(date, t, day_offset=0):
    """Combine date + time, applying day_offset days."""
    if date is None or t is None: return None
    base = date if isinstance(date, datetime.datetime) else datetime.datetime.combine(date, datetime.time())
    base = base + datetime.timedelta(days=day_offset)
    if isinstance(t, datetime.time):
        dt = datetime.datetime(base.year, base.month, base.day, t.hour, t.minute, t.second)
    elif isinstance(t, datetime.datetime):
        dt = datetime.datetime(base.year, base.month, base.day, t.hour, t.minute, t.second)
    else:
        return None
    return dt.strftime('%Y-%m-%dT%H:%M:%SZ')

def td_to_min(td):
    if td is None: return 0
    if isinstance(td, datetime.timedelta):
        return int(round(td.total_seconds() / 60))
    return 0

flights = []
auto_id = 1

# scan data rows
for r in range(HEADER_ROW + 1, ws.max_row + 1):
    date = ws.cell(r, C['date']).value if C['date'] else None
    flt  = ws.cell(r, C['flt']).value  if C['flt']  else None
    if date is None and flt is None:
        continue
    if not isinstance(date, (datetime.date, datetime.datetime)):
        continue

    off_t = ws.cell(r, C['off']).value if C['off'] else None
    ab_t  = ws.cell(r, C['ab']).value  if C['ab']  else None
    td_t  = ws.cell(r, C['td']).value  if C['td']  else None
    on_t  = ws.cell(r, C['on']).value  if C['on']  else None

    # day rollover detection — each successive time must be >= previous
    def mins(t):
        return t.hour*60 + t.minute if isinstance(t, datetime.time) else None
    off_m = mins(off_t)
    seq = [('off', off_t, 0)]
    last_m = off_m
    day = 0
    for tag, t in (('ab', ab_t), ('td', td_t), ('on', on_t)):
        m = mins(t)
        if m is not None and last_m is not None and m < last_m:
            day += 1
        seq.append((tag, t, day))
        if m is not None: last_m = m
    off_dt = to_iso_dt(date, off_t, 0)
    ab_dt  = to_iso_dt(date, ab_t,  seq[1][2])
    td_dt  = to_iso_dt(date, td_t,  seq[2][2])
    on_dt  = to_iso_dt(date, on_t,  seq[3][2])

    isLD   = ws.cell(r, C['isLD']).value   if C['isLD']   else None
    isPIC  = ws.cell(r, C['isPIC']).value  if C['isPIC']  else None
    role = 'PIC' if (isPIC == 1 or isPIC == 1.0) else 'SIC'
    # Duty inference: isL/D=1 ⇒ did the landing ⇒ PF. else PM (Cruise needs manual edit).
    duty = 'PF' if (isLD == 1 or isLD == 1.0) else 'PM'

    night_min = td_to_min(ws.cell(r, C['night']).value if C['night'] else None)
    ld_type   = ws.cell(r, C['ldType']).value if C['ldType'] else None
    if ld_type and isinstance(ld_type, str):
        ld_type = ld_type.strip()
        if ld_type not in ('Day','Night'): ld_type = None

    crew = []
    for k in ('crew1','crew2','crew3'):
        v = ws.cell(r, C[k]).value if C[k] else None
        crew.append('' if v is None else str(int(v)) if isinstance(v, float) else str(v))

    rem = ws.cell(r, C['remarks']).value if C['remarks'] else None

    flights.append({
        'id': auto_id,
        'dateUTC': date.strftime('%Y-%m-%d') if isinstance(date, (datetime.date, datetime.datetime)) else None,
        'fltNo': str(flt or '').strip(),
        'reg':   str(ws.cell(r, C['reg']).value or '').strip(),
        'acType':'A330',
        'dep':   str(ws.cell(r, C['dep']).value or '').strip().upper(),
        'dest':  str(ws.cell(r, C['dest']).value or '').strip().upper(),
        'offBlock':  off_dt,
        'airborne':  ab_dt,
        'touchdown': td_dt,
        'onBlock':   on_dt,
        'role': role,
        'duty': duty,
        'crew': crew,
        'remarks': str(rem) if rem else '',
        'nightTimeMin': night_min,
        'ldType': ld_type,
        'updatedAt': int(datetime.datetime.now().timestamp() * 1000),
    })
    auto_id += 1

# Aircraft = unique regs from the flights
regs = sorted({f['reg'] for f in flights if f['reg']})
aircraft = [{'reg': r, 'type': 'A330'} for r in regs]

out = {
    'version': 1,
    'exportedAt': datetime.datetime.now().isoformat() + 'Z',
    'flights': flights,
    'airports': airports,
    'aircraft': aircraft,
}
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)

print(f'Wrote {OUT}')
print(f'  Flights:  {len(flights)}')
print(f'  Airports: {len(airports)}')
print(f'  Aircraft: {len(aircraft)}')
